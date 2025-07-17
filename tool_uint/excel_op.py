import os
from datetime import datetime

from openpyxl import Workbook, load_workbook

now_time = datetime.now()
current_time = now_time.strftime("%Y-%m-%d")
class ExcelOp:
    def __init__(self):
        self.file_path=(r'C:\Users\MDPI\PycharmProjects\UI_Preprints\testdata\test_data.'
                        +current_time+'.xlsx')
        if os.path.exists(self.file_path):
            self.wb=load_workbook(self.file_path)
            self.ws = self.wb.active
        else:
            self.wb=Workbook()
            self.wb.save(self.file_path)
            self.wb=load_workbook(self.file_path)
            self.ws=self.wb.active
            self.ws.title="test_data"
            if self.ws.max_row == 1:
                self.ws.append(["preprint_id", "status", "time"])


    def add_data(self,data):
        self.ws.append(data)


    def get_data(self):
        data=[]
        for row in self.ws.iter_rows(min_row=2,values_only=True):
            data.append(row)
        return data

    def update_data(self,value,status):
        for row in self.ws.iter_rows(min_row=2,values_only=True):
            if row[0]==value:
                self.ws.cell(row=self.ws.max_row,column=2,value=status)
                self.ws.cell(row=self.ws.max_row,column=3,value=current_time)
                break

    def save_data(self):
        self.wb.save(self.file_path)

# excel_op=ExcelOp()
# excel_op=ExcelOp()
# get=excel_op.get_data()[-1][0]
# print(get)